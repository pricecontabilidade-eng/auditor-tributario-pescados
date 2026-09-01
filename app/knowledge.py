from __future__ import annotations
import re
from dataclasses import dataclass
from pathlib import Path
from datetime import date, datetime
import openpyxl

@dataclass
class NcmRule:
    ncm: str
    anexo: str
    cclasstrib: str | None
    permitido: bool
    ini: str | None
    fim: str | None

class KnowledgeBase:
    def __init__(self, base_dir: str | Path):
        self.base_dir = Path(base_dir)
        self.ncm_rules = self._load_ncm_rules()
        self.cclass = self._load_cclass()

    def _load_ncm_rules(self):
        p = self.base_dir / '10_NCM_ANEXOS.md'
        if not p.exists(): return []
        text = p.read_text(encoding='utf-8', errors='ignore')
        current_anexo = None
        rules=[]
        for line in text.splitlines():
            m = re.match(r'##\s+Anexo\s+([IVXLCDM]+)', line)
            if m:
                current_anexo=m.group(1)
                continue
            m = re.match(r'###\s+NCM\s+(\d{8})\s+—\s+(PERMITIDO|VEDADO)', line)
            if not m or not current_anexo: continue
            ncm, flag = m.group(1), m.group(2)
            # pull a small nearby block from text
            pos=text.find(line)
            block=text[pos:pos+900]
            cc=re.search(r'- \*\*cClassTrib:\*\*\s*([^\n]+)', block)
            ini=re.search(r'- \*\*Início da vigência:\*\*\s*([^\n]+)', block)
            fim=re.search(r'- \*\*Fim da vigência:\*\*\s*([^\n]+)', block)
            rules.append(NcmRule(ncm,current_anexo,cc.group(1).strip() if cc else None,flag=='PERMITIDO',ini.group(1).strip() if ini else None,fim.group(1).strip() if fim else None))
        return rules

    def _load_cclass(self):
        candidates=list(self.base_dir.glob('cClassTrib*.xlsx'))
        if not candidates: return {}
        wb=openpyxl.load_workbook(candidates[0], read_only=True, data_only=True)
        ws=wb[wb.sheetnames[0]]
        headers=[str(x.value).strip() if x.value is not None else '' for x in ws[1]]
        idx={h:i for i,h in enumerate(headers)}
        out={}
        for row in ws.iter_rows(min_row=2, values_only=True):
            cc=row[idx.get('cClassTrib',2)]
            if cc is None: continue
            cc=str(cc).strip().zfill(6)
            def g(k):
                j=idx.get(k); return row[j] if j is not None else None
            out[cc]={
                'cst': str(g('CST-IBS/CBS')).strip().zfill(3) if g('CST-IBS/CBS') is not None else None,
                'nome': g('Nome cClassTrib'), 'descricao':g('Descrição cClassTrib'), 'pred_ibs':g('pRedIBS'), 'pred_cbs':g('pRedCBS'),
                'ini':g('dIniVig'), 'fim':g('dFimVig'), 'anexo':g('ANEXO'), 'link':g('Link'), 'ind_nfe':g('indNFe')
            }
        return out

    def find_ncm(self,ncm):
        n=(ncm or '').replace('.','').strip().zfill(8)
        return [r for r in self.ncm_rules if r.ncm==n]

    def cclass_info(self, cc):
        if not cc: return None
        return self.cclass.get(str(cc).strip().zfill(6))
    def pis_cofins_pescado(self, ncm):
        """
        Retorna o tratamento esperado de PIS/COFINS para pescados
        conforme a regra documentada na Lei 10.925/2004, art. 1, XX,
        com redacao dada pela Lei 12.839/2013.

        Esta funcao nao presume tributacao normal para produtos
        que estejam fora do beneficio.
        """
        n = (ncm or "").replace(".", "").strip().zfill(8)

        # 03.02 - exceto 03029000
        if n.startswith("0302") and n != "03029000":
            return {
                "tratamento": "ALIQUOTA ZERO",
                "cst_pis": "06",
                "cst_cofins": "06",
                "aliq_pis": 0.0,
                "aliq_cofins": 0.0,
                "fundamento": (
                    "Lei 10.925/2004, art. 1, XX, "
                    "com redacao dada pela Lei 12.839/2013"
                ),
            }

        # 03.03 - peixes congelados
        if n.startswith("0303"):
            return {
                "tratamento": "ALIQUOTA ZERO",
                "cst_pis": "06",
                "cst_cofins": "06",
                "aliq_pis": 0.0,
                "aliq_cofins": 0.0,
                "fundamento": (
                    "Lei 10.925/2004, art. 1, XX, "
                    "com redacao dada pela Lei 12.839/2013"
                ),
            }

        # 03.04 - files e outras carnes de peixes
        if n.startswith("0304"):
            return {
                "tratamento": "ALIQUOTA ZERO",
                "cst_pis": "06",
                "cst_cofins": "06",
                "aliq_pis": 0.0,
                "aliq_cofins": 0.0,
                "fundamento": (
                    "Lei 10.925/2004, art. 1, XX, "
                    "com redacao dada pela Lei 12.839/2013"
                ),
            }

        # Excecao expressa da posicao 03.02
        if n == "03029000":
            return {
                "tratamento": "VALIDAR OUTRA BASE LEGAL",
                "cst_pis": None,
                "cst_cofins": None,
                "aliq_pis": None,
                "aliq_cofins": None,
                "fundamento": (
                    "NCM 03029000 excluido da regra especifica "
                    "da Lei 12.839/2013"
                ),
            }

        # 03.06 - crustaceos
        if n.startswith("0306"):
            return {
                "tratamento": "VALIDAR OUTRA BASE LEGAL",
                "cst_pis": None,
                "cst_cofins": None,
                "aliq_pis": None,
                "aliq_cofins": None,
                "fundamento": (
                    "Crustaceo nao abrangido automaticamente pela "
                    "regra especifica da Lei 12.839/2013"
                ),
            }

        # 03.07 - moluscos
        if n.startswith("0307"):
            return {
                "tratamento": "VALIDAR OUTRA BASE LEGAL",
                "cst_pis": None,
                "cst_cofins": None,
                "aliq_pis": None,
                "aliq_cofins": None,
                "fundamento": (
                    "Molusco nao abrangido automaticamente pela "
                    "regra especifica da Lei 12.839/2013"
                ),
            }

        return None
