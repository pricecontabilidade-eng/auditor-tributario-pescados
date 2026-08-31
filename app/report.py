from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def dataframe_to_xlsx(df: pd.DataFrame) -> bytes:
    bio=BytesIO()
    with pd.ExcelWriter(bio, engine='openpyxl') as writer:
        df.to_excel(writer,index=False,sheet_name='Auditoria')
        ws=writer.book['Auditoria']
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        fill=PatternFill('solid', fgColor='1F4E78')
        for c in ws[1]:
            c.font=Font(color='FFFFFF',bold=True); c.fill=fill; c.alignment=Alignment(wrap_text=True)
        for i,col in enumerate(df.columns,1):
            maxlen=max([len(str(col))]+[len(str(v)) for v in df[col].head(200).fillna('')])
            ws.column_dimensions[get_column_letter(i)].width=min(max(maxlen+2,12),45)
        for row in ws.iter_rows(min_row=2):
            for c in row: c.alignment=Alignment(vertical='top',wrap_text=True)
                        # Destaque visual da coluna Status
        if "Status" in df.columns:
            status_col = df.columns.get_loc("Status") + 1

            fill_correto = PatternFill(
                fill_type="solid",
                fgColor="C6EFCE"
            )

            fill_divergente = PatternFill(
                fill_type="solid",
                fgColor="FFC7CE"
            )

            fill_pendente = PatternFill(
                fill_type="solid",
                fgColor="FFEB9C"
            )

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=status_col)
                status = str(cell.value or "").strip().upper()

                if status == "CORRETO":
                    cell.fill = fill_correto
                    cell.font = Font(bold=True)

                elif status == "DIVERGENTE":
                    cell.fill = fill_divergente
                    cell.font = Font(bold=True)

                elif status == "PENDENTE DE VALIDAÇÃO":
                    cell.fill = fill_pendente
                    cell.font = Font(bold=True)
    return bio.getvalue()
